"""Report writer tool for generating formatted documents"""
import os
from datetime import datetime
from langchain_core.tools import tool


@tool
def write_file(
    title: str,
    content: str,
    report_type: str = "general",
    format: str = "text",
    filename: str = None
) -> str:
    """Generate a professional formatted report.

    Use this tool to create structured reports, analyses, and documents
    based on data gathered from other tools.

    Args:
        title: Report title
        content: Main content/analysis for the report (can include sections, bullet points, etc.)
            For Excel format, content should be structured data:
            - Simple text will create a single-column report
            - Use "||" to separate columns in each row
            - Use "\\n" to separate rows
            Example for Excel: "Produit||Prix||Stock\\nE-City||1299||45\\nUrban-X||899||120"
        report_type: Type of report structure:
            - "general": Standard report (Introduction, Content, Conclusion)
            - "technical": Technical analysis (Résumé, Analyse Technique, Problèmes, Solutions)
            - "business": Business report (Contexte, Analyse, Opportunités, Recommandations)
            - "customer_analysis": Customer analysis (Retours Clients, Points Positifs, Améliorations)
        format: Output format:
            - "text": Plain text (default)
            - "excel": Excel spreadsheet (.xlsx)
        filename: Output filename for Excel format (without extension).
            If not provided for Excel, uses sanitized title.
            File will be saved in current directory.

    Returns:
        For text format: Formatted report as string
        For Excel: Path to the generated .xlsx file

    Examples:
        write_file("Analyse des Ventes Q4", "Résumé: Ventes en hausse de 15%...", report_type="business")
        write_file("Rapport Technique E-City", content, report_type="technical")
        write_file("Stock Report", "Produit||Prix||Stock\\nE-City||1299||45", format="excel")
    """
    try:
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        if format == "excel":
            return _generate_excel(title, content, report_type, timestamp, filename)
        else:
            return _generate_text(title, content, report_type, timestamp)

    except Exception as e:
        return f"Error generating report: {str(e)}"


def _generate_text(title: str, content: str, report_type: str, timestamp: str) -> str:
    """Generate plain text report"""
    return f"""
{'='*60}
{title.upper()}
{'='*60}

Type: {report_type.replace('_', ' ').title()}
Date: {timestamp}

{'-'*60}

{content}

{'-'*60}
Fin du rapport
{'='*60}
"""


def _generate_excel(title: str, content: str, report_type: str, timestamp: str, filename: str = None) -> str:
    """Generate Excel (.xlsx) report using openpyxl"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return "Error: openpyxl is not installed. Install it with: pip install openpyxl"

    wb = Workbook()
    ws = wb.active
    ws.title = "Rapport"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=14)
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    meta_fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title row
    ws.merge_cells('A1:E1')
    ws['A1'] = title
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Metadata rows
    ws['A3'] = "Type:"
    ws['B3'] = report_type.replace('_', ' ').title()
    ws['A4'] = "Date:"
    ws['B4'] = timestamp
    ws['A5'] = "Généré par:"
    ws['B5'] = "Agent RAG VéloCorp"

    for row in range(3, 6):
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].fill = meta_fill
        ws[f'B{row}'].fill = meta_fill

    # Content - parse structured data
    start_row = 7
    lines = content.strip().split('\n')

    for i, line in enumerate(lines):
        row_num = start_row + i
        if '||' in line:
            # Structured data with columns
            cells = line.split('||')
            for j, cell in enumerate(cells):
                col_letter = chr(ord('A') + j)
                ws[f'{col_letter}{row_num}'] = cell.strip()
                ws[f'{col_letter}{row_num}'].border = thin_border
                # Bold first row (header)
                if i == 0:
                    ws[f'{col_letter}{row_num}'].font = Font(bold=True)
                    ws[f'{col_letter}{row_num}'].fill = PatternFill(
                        start_color="3498DB", end_color="3498DB", fill_type="solid"
                    )
                    ws[f'{col_letter}{row_num}'].font = Font(bold=True, color="FFFFFF")
        else:
            # Simple text content
            ws[f'A{row_num}'] = line

    # Auto-adjust column widths (skip merged cells)
    for column in ws.columns:
        max_length = 0
        column_letter = None
        for cell in column:
            try:
                # Skip MergedCell objects
                if hasattr(cell, 'column_letter'):
                    if column_letter is None:
                        column_letter = cell.column_letter
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        if column_letter:
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    # Generate filename
    if not filename:
        # Sanitize title for filename
        filename = "".join(c if c.isalnum() or c in ' -_' else '_' for c in title)
        filename = filename.replace(' ', '_')

    filepath = f"{filename}.xlsx"

    wb.save(filepath)

    return f"Excel report generated: {os.path.abspath(filepath)}"
